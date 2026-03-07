import argparse
import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


SHEETS = ["Summary", "LineItems", "Assumptions", "Evidence"]

LINE_ITEMS_COLUMN_MAP = {
    "resource_type": ["resource_type", "service"],
    "system": ["system"],
    "env": ["env", "environment"],
    "SAP_workload": ["SAP_workload", "sap_workload"],
    "workload_type": ["workload_type"],
    "vcpu": ["vcpu", "parsed_vcpu"],
    "memory_gb": ["memory_gb", "parsed_memory_gb"],
    "disk": ["disk"],
    "quantity": ["quantity"],
    "os": ["os"],
    "region": ["region", "region_input"],
    "item_id": ["item_id", "line_id"],
    "provider": ["provider"],
    "sku/os": ["sku/os", "sku_os", "os", "sku"],
    "region_azure": ["region_azure"],
    "primary_sku": ["primary_sku", "sku"],
    "fallback_sku": ["fallback_sku", "fallback_skus"],
    "sap_sku": ["sap_sku"],
    "billing_unit": ["billing_unit", "unit"],
    "AWS_paygo": ["AWS_paygo", "unit_price_AWS_paygo", "aws_paygo_hourly_usd"],
    "AWS_1YRI": ["AWS_1YRI", "aws_ri_1y_hourly_usd"],
    "AWS_3YRI": ["AWS_3YRI", "aws_ri_3y_hourly_usd"],
    "Azure_paygo": ["Azure_paygo", "unit_price_Azure_paygo", "azure_paygo_hourly_usd", "unit_price_hourly"],
    "Azure_1YRI": ["Azure_1YRI", "azure_ri_1y_hourly_usd"],
    "Azure_3YRI": ["Azure_3YRI", "azure_ri_3y_hourly_usd"],
    "Azure_SAP_paygo": ["Azure_SAP_paygo", "unit_price_Azure_SAP_paygo", "sap_azure_paygo_hourly_usd"],
    "Azure_SAP_1YRI": ["Azure_SAP_1YRI", "sap_azure_ri_1y_hourly_usd"],
    "Azure_SAP_3YRI": ["Azure_SAP_3YRI", "sap_azure_ri_3y_hourly_usd"],
    "review_flag": ["review_flag"],
    "review_reason": ["review_reason", "notes"],
    "evidence_id": ["evidence_id"],
}

LINE_ITEMS_BASE_HEADERS = [
    "resource_type",
    "system",
    "env",
    "SAP_workload",
    "workload_type",
    "vcpu",
    "memory_gb",
    "disk",
    "quantity",
    "os",
    "region",
    "item_id",
    "provider",
    "sku/os",
    "region_azure",
    "primary_sku",
    "fallback_sku",
    "sap_sku",
    "billing_unit",
    "AWS_paygo",
    "AWS_1YRI",
    "AWS_3YRI",
    "Azure_paygo",
    "Azure_1YRI",
    "Azure_3YRI",
    "Azure_SAP_paygo",
    "Azure_SAP_1YRI",
    "Azure_SAP_3YRI",
    "review_flag",
    "review_reason",
    "evidence_id",
]

LINE_ITEMS_AWS_COLUMNS_WHEN_NO_PROVIDER = {
    "provider",
    "AWS_paygo",
    "AWS_1YRI",
    "AWS_3YRI",
}


def _normalize_key(key: str) -> str:
    return "".join(ch for ch in str(key or "").strip().lower() if ch.isalnum() or ch == "_")


def _coalesce_value(obj: Dict[str, Any], keys: list[str], default: Any = None) -> Any:
    key_map = {_normalize_key(k): k for k in obj.keys()}
    for key in keys:
        norm = _normalize_key(key)
        if norm in key_map:
            return obj[key_map[norm]]
    return default


def _to_float(value: Any) -> float | None:
    try:
        if value is None or str(value).strip() == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _sum_or_none(values: list[float | None]) -> float | None:
    filtered = [value for value in values if value is not None]
    if not filtered:
        return None
    return float(sum(filtered))


def _safe_divide(num: float | None, den: float | None) -> float | None:
    if num is None or den in (None, 0):
        return None
    return num / den


def _first_non_empty(values: list[Any], default: Any = "") -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return default


def _print(obj: Dict[str, Any]) -> None:
    print(json.dumps(obj, ensure_ascii=False, indent=2))


def _as_bool(value: Any, default: bool = True) -> bool:
    if value is None:
        return default
    token = str(value).strip().lower()
    if token in {"1", "true", "yes", "y", "on"}:
        return True
    if token in {"0", "false", "no", "n", "off"}:
        return False
    return default


def init_template(template_path: Path) -> None:
    template_path.parent.mkdir(parents=True, exist_ok=True)

    wb = _build_template_workbook()
    wb.save(template_path)


def _build_template_workbook() -> Workbook:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_line = wb.create_sheet("LineItems")
    ws_ass = wb.create_sheet("Assumptions")
    ws_evd = wb.create_sheet("Evidence")

    ws_summary.merge_cells("A1:C1")
    ws_summary["A1"] = "Quote Summary"
    ws_summary["A2"] = "Field"
    ws_summary["B2"] = "Value"
    ws_summary["C2"] = "Notes"
    ws_summary["A3"] = "Customer / Project"
    ws_summary["A4"] = "Region"
    ws_summary["A5"] = "Currency"
    ws_summary["A6"] = "Competitor Cloud"
    ws_summary["A7"] = "Pricing Source Date"
    ws_summary.merge_cells("A9:D9")
    ws_summary["A9"] = "Hourly Price Comparison (USD)"
    ws_summary["B10"] = "PayGo"
    ws_summary["C10"] = "1Y RI"
    ws_summary["D10"] = "3Y RI"
    ws_summary["A11"] = "Competitor (Original)"
    ws_summary["A12"] = "Azure (Mapped)"
    ws_summary["A13"] = "Delta (Azure - Comp.)"
    ws_summary["A14"] = "Delta %"

    for row in (1, 2, 9, 10):
        for cell in ws_summary[row]:
            cell.font = Font(bold=True)
    for row in range(1, 15):
        for col in ("A", "B", "C", "D"):
            ws_summary[f"{col}{row}"].alignment = Alignment(vertical="center")

    line_headers = [
        "item_id",
        "provider",
        "resource_type",
        "quantity",
        "sku/os",
        "region",
        "region_azure",
        "primary_sku",
        "fallback_sku",
        "sap_sku",
        "billing_unit",
        "AWS_paygo",
        "AWS_1YRI",
        "AWS_3YRI",
        "Azure_paygo",
        "Azure_1YRI",
        "Azure_3YRI",
        "Azure_SAP_paygo",
        "Azure_SAP_1YRI",
        "Azure_SAP_3YRI",
        "review_flag",
        "review_reason",
        "evidence_id",
    ]
    for idx, header in enumerate(line_headers, start=1):
        ws_line.cell(1, idx, header)

    ass_headers = ["assumption_id", "category", "statement", "impact_scope", "requires_confirmation"]
    for idx, header in enumerate(ass_headers, start=1):
        ws_ass.cell(1, idx, header)

    evd_headers = [
        "evidence_id",
        "item_id",
        "source_type",
        "source_url",
        "fetched_at",
        "mapping_version",
        "policy_version",
        "kb_version",
        "price_date",
        "source_ref",
        "status",
    ]
    for idx, header in enumerate(evd_headers, start=1):
        ws_evd.cell(1, idx, header)

    return wb


def load_payload(input_json: Path) -> Dict[str, Any]:
    with input_json.open("r", encoding="utf-8") as file:
        payload = json.load(file)

    if "summary" not in payload or "line_items" not in payload:
        raise ValueError("input json must contain `summary` and `line_items`")
    return payload


def _sheet_headers(ws) -> list[str]:
    headers: list[str] = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        headers.append(str(value).strip() if value is not None else "")
    return headers


def _clear_data_rows(ws, header_row: int = 1) -> None:
    if ws.max_row <= header_row:
        return
    ws.delete_rows(header_row + 1, ws.max_row - header_row)


def _write_table_rows(ws, headers: list[str], rows: list[Dict[str, Any]]) -> int:
    _clear_data_rows(ws, header_row=1)
    count = 0
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, row_data.get(header))
        count += 1
    return count


def _normalize_line_items(line_items: List[Dict[str, Any]]) -> list[Dict[str, Any]]:
    normalized_rows: list[Dict[str, Any]] = []
    for idx, item in enumerate(line_items, start=1):
        row: dict[str, Any] = {}
        for header, candidates in LINE_ITEMS_COLUMN_MAP.items():
            row[header] = _coalesce_value(item, candidates)

        if row.get("item_id") in (None, ""):
            row["item_id"] = f"item_{idx}"

        if row.get("sku/os") in (None, ""):
            sku = _coalesce_value(item, ["sku", "primary_sku"])
            os_name = _coalesce_value(item, ["os"])
            row["sku/os"] = _first_non_empty([os_name, sku], "")

        normalized_rows.append(row)

    return normalized_rows


def _canonicalize_line_sheet_headers(ws) -> None:
    header_renames = {
        "monthly_cost_AWS_paygo": "line_total_AWS_paygo",
        "monthly_cost_Azure_paygo": "line_total_Azure_paygo",
        "monthly_cost_AWS_1YRI": "line_total_AWS_1YRI",
        "monthly_cost_AWS_3YRI": "line_total_AWS_3YRI",
        "monthly_cost_Azure_1YRI": "line_total_Azure_1YRI",
        "monthly_cost_Azure_3YRI": "line_total_Azure_3YRI",
    }
    required_headers = [
        "item_id",
        "provider",
        "resource_type",
        "quantity",
        "sku/os",
        "region",
        "region_azure",
        "primary_sku",
        "fallback_sku",
        "sap_sku",
        "billing_unit",
        "AWS_paygo",
        "AWS_1YRI",
        "AWS_3YRI",
        "Azure_paygo",
        "Azure_1YRI",
        "Azure_3YRI",
        "Azure_SAP_paygo",
        "Azure_SAP_1YRI",
        "Azure_SAP_3YRI",
        "review_flag",
        "review_reason",
        "evidence_id",
    ]
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value in header_renames:
            ws.cell(1, col).value = header_renames[value]

    deprecated_headers = {
        "unit_price_AWS_paygo",
        "unit_price_Azure_paygo",
        "unit_price_Azure_SAP_paygo",
        "line_total_AWS_paygo",
        "line_total_Azure_paygo",
        "line_total_AWS_1YRI",
        "line_total_AWS_3YRI",
        "line_total_Azure_1YRI",
        "line_total_Azure_3YRI",
        "line_total_Azure_SAP_paygo",
        "line_total_Azure_SAP_1YRI",
        "line_total_Azure_SAP_3YRI",
        "fallback_skus",
    }
    _drop_line_sheet_headers(ws, deprecated_headers)

    existing = set(_sheet_headers(ws))
    next_col = ws.max_column + 1
    for header in required_headers:
        if header in existing:
            continue
        ws.cell(1, next_col).value = header
        next_col += 1


def _drop_line_sheet_headers(ws, headers_to_drop: set[str]) -> None:
    for col in range(ws.max_column, 0, -1):
        value = ws.cell(1, col).value
        if value in headers_to_drop:
            ws.delete_cols(col, 1)


def _set_line_sheet_headers(ws, headers: list[str]) -> None:
    current_max = ws.max_column
    if current_max > len(headers):
        ws.delete_cols(len(headers) + 1, current_max - len(headers))
    elif current_max < len(headers):
        ws.insert_cols(current_max + 1, len(headers) - current_max)

    for idx, header in enumerate(headers, start=1):
        ws.cell(1, idx).value = header


def _normalize_assumptions(assumptions: Any) -> list[Dict[str, Any]]:
    if assumptions is None:
        return []
    if isinstance(assumptions, dict):
        rows = []
        for idx, (key, value) in enumerate(assumptions.items(), start=1):
            rows.append(
                {
                    "assumption_id": f"A-{idx}",
                    "category": "mapping",
                    "statement": f"{key}={value}",
                    "impact_scope": "subset rows",
                    "requires_confirmation": False,
                }
            )
        return rows
    if isinstance(assumptions, list):
        rows: list[dict[str, Any]] = []
        for idx, item in enumerate(assumptions, start=1):
            if isinstance(item, dict):
                rows.append(
                    {
                        "assumption_id": _coalesce_value(item, ["assumption_id"], f"A-{idx}"),
                        "category": _coalesce_value(item, ["category", "source"], "mapping"),
                        "statement": _coalesce_value(item, ["statement", "value", "key"], ""),
                        "impact_scope": _coalesce_value(item, ["impact_scope", "notes"], "subset rows"),
                        "requires_confirmation": _coalesce_value(item, ["requires_confirmation"], False),
                    }
                )
            else:
                rows.append(
                    {
                        "assumption_id": f"A-{idx}",
                        "category": "mapping",
                        "statement": str(item),
                        "impact_scope": "subset rows",
                        "requires_confirmation": False,
                    }
                )
        return rows
    raise ValueError("assumptions must be dict or list")


def _normalize_evidence(evidence: Any, normalized_line_items: list[Dict[str, Any]]) -> list[Dict[str, Any]]:
    if evidence is None:
        return []
    if not isinstance(evidence, list):
        raise ValueError("evidence must be list")

    rows: list[dict[str, Any]] = []
    for idx, item in enumerate(evidence, start=1):
        if not isinstance(item, dict):
            continue
        item_id_default = normalized_line_items[idx - 1]["item_id"] if idx - 1 < len(normalized_line_items) else None
        rows.append(
            {
                "evidence_id": _coalesce_value(item, ["evidence_id"], f"ev-{idx}"),
                "item_id": _coalesce_value(item, ["item_id", "line_id"], item_id_default),
                "source_type": _coalesce_value(item, ["source_type", "evidence_type"], "retail_api"),
                "source_url": _coalesce_value(item, ["source_url", "ref"], ""),
                "fetched_at": _coalesce_value(item, ["fetched_at"], ""),
                "mapping_version": _coalesce_value(item, ["mapping_version"], "v1"),
                "policy_version": _coalesce_value(item, ["policy_version"], "v1"),
                "kb_version": _coalesce_value(item, ["kb_version"], "v1"),
                "price_date": _coalesce_value(item, ["price_date", "fetched_at"], ""),
                "source_ref": _coalesce_value(item, ["source_ref", "detail"], ""),
                "status": _coalesce_value(item, ["status"], "ok"),
            }
        )
    return rows


def _build_summary_cells(summary: Dict[str, Any], normalized_line_items: list[Dict[str, Any]]) -> dict[str, Any]:
    source = {_normalize_key(k): v for k, v in summary.items()}

    def get_value(*keys: str, default: Any = None) -> Any:
        for key in keys:
            norm = _normalize_key(key)
            if norm in source:
                return source[norm]
        return default

    c_paygo = _to_float(get_value("competitor_paygo", "c_paygo", "competitor_original_paygo"))
    c_1y = _to_float(get_value("competitor_1y_ri", "c_1y", "competitor_original_1yri"))
    c_3y = _to_float(get_value("competitor_3y_ri", "c_3y", "competitor_original_3yri"))
    a_paygo = _to_float(get_value("azure_paygo", "a_paygo", "azure_mapped_paygo"))
    a_1y = _to_float(get_value("azure_1y_ri", "a_1y", "azure_mapped_1yri"))
    a_3y = _to_float(get_value("azure_3y_ri", "a_3y", "azure_mapped_3yri"))

    if any(value is None for value in [c_paygo, c_1y, c_3y, a_paygo, a_1y, a_3y]):
        c_paygo_values = [_to_float(item.get("AWS_paygo")) for item in normalized_line_items]
        c_1y_values = [_to_float(item.get("AWS_1YRI")) for item in normalized_line_items]
        c_3y_values = [_to_float(item.get("AWS_3YRI")) for item in normalized_line_items]
        a_paygo_values = [_to_float(item.get("Azure_paygo")) for item in normalized_line_items]
        a_1y_values = [_to_float(item.get("Azure_1YRI")) for item in normalized_line_items]
        a_3y_values = [_to_float(item.get("Azure_3YRI")) for item in normalized_line_items]

        c_paygo = c_paygo if c_paygo is not None else _sum_or_none(c_paygo_values)
        c_1y = c_1y if c_1y is not None else _sum_or_none(c_1y_values)
        c_3y = c_3y if c_3y is not None else _sum_or_none(c_3y_values)
        a_paygo = a_paygo if a_paygo is not None else _sum_or_none(a_paygo_values)
        a_1y = a_1y if a_1y is not None else _sum_or_none(a_1y_values)
        a_3y = a_3y if a_3y is not None else _sum_or_none(a_3y_values)

    d_paygo = (a_paygo - c_paygo) if a_paygo is not None and c_paygo is not None else None
    d_1y = (a_1y - c_1y) if a_1y is not None and c_1y is not None else None
    d_3y = (a_3y - c_3y) if a_3y is not None and c_3y is not None else None

    dp_paygo = _safe_divide(d_paygo, c_paygo)
    dp_1y = _safe_divide(d_1y, c_1y)
    dp_3y = _safe_divide(d_3y, c_3y)

    pricing_source_note = get_value("pricing_source_note", default="Azure Retail Prices API")

    return {
        "B3": get_value("customer_project", "customer", default="N/A"),
        "B4": get_value("region", default="N/A"),
        "C4": get_value("region_note", default=""),
        "B5": get_value("currency", default="USD (excl. tax)"),
        "B6": get_value("competitor_cloud", default="N/A"),
        "B7": get_value("pricing_source_date", default=""),
        "C7": pricing_source_note,
        "B11": c_paygo,
        "C11": c_1y,
        "D11": c_3y,
        "B12": a_paygo,
        "C12": a_1y,
        "D12": a_3y,
        "B13": d_paygo,
        "C13": d_1y,
        "D13": d_3y,
        "B14": dp_paygo,
        "C14": dp_1y,
        "D14": dp_3y,
    }


def _apply_summary_number_formats(ws) -> None:
    for row in range(11, 14):
        for col in ("B", "C", "D"):
            ws[f"{col}{row}"].number_format = '#,##0.00'
    for col in ("B", "C", "D"):
        ws[f"{col}14"].number_format = '0.00%'


def write_quote_excel(payload: Dict[str, Any], output_xlsx: Path) -> Dict[str, Any]:
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    workbook = _build_template_workbook()

    summary_sheet = workbook["Summary"]
    line_sheet = workbook["LineItems"]
    assumptions_sheet = workbook["Assumptions"]
    evidence_sheet = workbook["Evidence"]

    include_provider_columns = _as_bool(payload.get("summary", {}).get("input_provider_present"), default=True)
    line_headers_order = list(LINE_ITEMS_BASE_HEADERS)
    if not include_provider_columns:
        line_headers_order = [header for header in line_headers_order if header not in LINE_ITEMS_AWS_COLUMNS_WHEN_NO_PROVIDER]
    _set_line_sheet_headers(line_sheet, line_headers_order)

    line_items = _normalize_line_items(payload.get("line_items", []))
    assumptions = _normalize_assumptions(payload.get("assumptions"))
    evidence = _normalize_evidence(payload.get("evidence"), line_items)
    summary_cells = _build_summary_cells(payload.get("summary", {}), line_items)

    for cell, value in summary_cells.items():
        summary_sheet[cell] = value
    summary_sheet["A9"] = "Hourly Price Comparison (USD)"
    if summary_sheet["C7"].value in (None, ""):
        summary_sheet["C7"] = "Azure Retail Prices API"
    _apply_summary_number_formats(summary_sheet)

    line_headers = _sheet_headers(line_sheet)
    assumption_headers = _sheet_headers(assumptions_sheet)
    evidence_headers = _sheet_headers(evidence_sheet)

    line_count = _write_table_rows(line_sheet, line_headers, line_items)
    assumption_count = _write_table_rows(assumptions_sheet, assumption_headers, assumptions)
    evidence_count = _write_table_rows(evidence_sheet, evidence_headers, evidence)

    workbook.save(output_xlsx)

    return {
        "status": "ok",
        "output_file": str(output_xlsx),
        "sheets_written": SHEETS,
        "row_counts": {
            "Summary": 12,
            "LineItems": line_count,
            "Assumptions": assumption_count,
            "Evidence": evidence_count,
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Write structured quote payload into Excel workbook.")
    parser.add_argument("--input-json", help="Path to structured quote JSON payload.")
    parser.add_argument("--output-xlsx", help="Output Excel path.")
    parser.add_argument("--init-template", action="store_true", help="Initialize template workbook and exit.")
    args = parser.parse_args()

    try:
        if args.init_template:
            if not args.output_xlsx:
                raise ValueError("`--output-xlsx` is required when `--init-template` is used")
            init_template(Path(args.output_xlsx))
            _print(
                {
                    "status": "ok",
                    "message": "template initialized",
                    "template_file": str(Path(args.output_xlsx)),
                }
            )
            return

        if not args.input_json or not args.output_xlsx:
            raise ValueError("`--input-json` and `--output-xlsx` are required")

        payload = load_payload(Path(args.input_json))
        result = write_quote_excel(payload, Path(args.output_xlsx))
        _print(result)

    except Exception as exc:  # noqa: BLE001
        _print({"status": "error", "error": str(exc)})


if __name__ == "__main__":
    main()
